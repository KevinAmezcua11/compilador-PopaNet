# excel_export.py
import openpyxl
from tkinter import filedialog, messagebox

def export_to_excel(vlsm_data):
    """
    Exporta los resultados del cálculo VLSM a un archivo de Excel (.xlsx).
    Cada red se guarda en una hoja diferente.
    """
    if not vlsm_data:
        messagebox.showerror("Error", "No hay datos para exportar.")
        return

    # Solicita al usuario la ruta donde guardar el archivo
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")],
        title="Guardar resultados de VLSM"
    )

    if not file_path:
        return  # Usuario canceló

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # Agrupa las subredes por nombre de red o IP base
    grouped = {}
    for subred in vlsm_data:
        group_key = subred.get("nombre_red") or subred["ip_base"]
        if group_key not in grouped:
            grouped[group_key] = []
        grouped[group_key].append(subred)

    for nombre_red, subredes in grouped.items():
        # Crea una hoja por cada red
        sheet = workbook.create_sheet(title=f"Red {nombre_red}")

        headers = ["Subred"] + [k for k in subredes[0].keys() if k not in ["ip_base", "nombre_red"]]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        for row_num, subred in enumerate(subredes, start=2):
            sheet.cell(row=row_num, column=1, value=row_num - 1)
            col_index = 2
            for key in headers[1:]:
                sheet.cell(row=row_num, column=col_index, value=subred[key])
                col_index += 1

    workbook.save(file_path)
    messagebox.showinfo("Exportar a Excel", "Datos exportados exitosamente.")
